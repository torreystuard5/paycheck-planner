import csv
import io
import logging
from decimal import Decimal, InvalidOperation
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, numbers

logger = logging.getLogger(__name__)


# ── Excel exports ───────────────────────────────────────────────────


def _currency_format(ws, col_letter, min_row, max_row):
    """Apply currency number format to a column range."""
    for row in range(min_row, max_row + 1):
        cell = ws[f"{col_letter}{row}"]
        cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1


def _header_style(ws, headers):
    """Bold the header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)


def export_bills_excel(bills: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bills"
    headers = ["Name", "Amount", "Due Day", "Frequency", "Category", "Auto Pay", "Reminder Days"]
    _header_style(ws, headers)

    for i, bill in enumerate(bills, 2):
        ws.cell(row=i, column=1, value=bill.name)
        ws.cell(row=i, column=2, value=float(bill.amount))
        ws.cell(row=i, column=3, value=bill.due_day)
        ws.cell(row=i, column=4, value=bill.frequency)
        ws.cell(row=i, column=5, value=bill.category or "")
        ws.cell(row=i, column=6, value="Yes" if bill.auto_pay else "No")
        ws.cell(row=i, column=7, value=bill.reminder_days)

    if bills:
        _currency_format(ws, "B", 2, len(bills) + 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_debts_excel(debts: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Debts"
    headers = ["Name", "Type", "Balance", "Credit Limit", "APR%", "Minimum Payment", "Due Day", "Auto Pay"]
    _header_style(ws, headers)

    for i, debt in enumerate(debts, 2):
        ws.cell(row=i, column=1, value=debt.name)
        ws.cell(row=i, column=2, value=debt.type)
        ws.cell(row=i, column=3, value=float(debt.balance))
        ws.cell(row=i, column=4, value=float(debt.credit_limit) if debt.credit_limit else "")
        ws.cell(row=i, column=5, value=float(debt.apr))
        ws.cell(row=i, column=6, value=float(debt.minimum_payment))
        ws.cell(row=i, column=7, value=debt.due_day)
        ws.cell(row=i, column=8, value="Yes" if debt.auto_pay else "No")

    if debts:
        _currency_format(ws, "C", 2, len(debts) + 1)
        _currency_format(ws, "F", 2, len(debts) + 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_payments_excel(payments: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    headers = ["Date Paid", "Bill/Debt Name", "Amount", "Pay Period", "Extra Payment"]
    _header_style(ws, headers)

    for i, p in enumerate(payments, 2):
        ws.cell(row=i, column=1, value=str(p.paid_date) if p.paid_date else "")
        # Try to get related name
        name = ""
        if hasattr(p, "bill") and p.bill:
            name = p.bill.name
        elif hasattr(p, "debt") and p.debt:
            name = p.debt.name
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=float(p.amount))
        ws.cell(row=i, column=4, value=str(p.pay_period_date) if p.pay_period_date else "")
        ws.cell(row=i, column=5, value="Yes" if p.is_extra else "No")

    if payments:
        _currency_format(ws, "C", 2, len(payments) + 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_all_excel(bills: list, debts: list, payments: list) -> io.BytesIO:
    wb = Workbook()

    # Bills sheet
    ws_bills = wb.active
    ws_bills.title = "Bills"
    bill_headers = ["Name", "Amount", "Due Day", "Frequency", "Category", "Auto Pay", "Reminder Days"]
    _header_style(ws_bills, bill_headers)
    for i, bill in enumerate(bills, 2):
        ws_bills.cell(row=i, column=1, value=bill.name)
        ws_bills.cell(row=i, column=2, value=float(bill.amount))
        ws_bills.cell(row=i, column=3, value=bill.due_day)
        ws_bills.cell(row=i, column=4, value=bill.frequency)
        ws_bills.cell(row=i, column=5, value=bill.category or "")
        ws_bills.cell(row=i, column=6, value="Yes" if bill.auto_pay else "No")
        ws_bills.cell(row=i, column=7, value=bill.reminder_days)
    if bills:
        _currency_format(ws_bills, "B", 2, len(bills) + 1)

    # Debts sheet
    ws_debts = wb.create_sheet("Debts")
    debt_headers = ["Name", "Type", "Balance", "Credit Limit", "APR%", "Minimum Payment", "Due Day", "Auto Pay"]
    _header_style(ws_debts, debt_headers)
    for i, debt in enumerate(debts, 2):
        ws_debts.cell(row=i, column=1, value=debt.name)
        ws_debts.cell(row=i, column=2, value=debt.type)
        ws_debts.cell(row=i, column=3, value=float(debt.balance))
        ws_debts.cell(row=i, column=4, value=float(debt.credit_limit) if debt.credit_limit else "")
        ws_debts.cell(row=i, column=5, value=float(debt.apr))
        ws_debts.cell(row=i, column=6, value=float(debt.minimum_payment))
        ws_debts.cell(row=i, column=7, value=debt.due_day)
        ws_debts.cell(row=i, column=8, value="Yes" if debt.auto_pay else "No")
    if debts:
        _currency_format(ws_debts, "C", 2, len(debts) + 1)
        _currency_format(ws_debts, "F", 2, len(debts) + 1)

    # Payments sheet
    ws_pay = wb.create_sheet("Payments")
    pay_headers = ["Date Paid", "Bill/Debt Name", "Amount", "Pay Period", "Extra Payment"]
    _header_style(ws_pay, pay_headers)
    for i, p in enumerate(payments, 2):
        ws_pay.cell(row=i, column=1, value=str(p.paid_date) if p.paid_date else "")
        name = ""
        if hasattr(p, "bill") and p.bill:
            name = p.bill.name
        elif hasattr(p, "debt") and p.debt:
            name = p.debt.name
        ws_pay.cell(row=i, column=2, value=name)
        ws_pay.cell(row=i, column=3, value=float(p.amount))
        ws_pay.cell(row=i, column=4, value=str(p.pay_period_date) if p.pay_period_date else "")
        ws_pay.cell(row=i, column=5, value="Yes" if p.is_extra else "No")
    if payments:
        _currency_format(ws_pay, "C", 2, len(payments) + 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── CSV exports ─────────────────────────────────────────────────────


def export_bills_csv(bills: list) -> io.BytesIO:
    buf = io.BytesIO()
    wrapper = io.TextIOWrapper(buf, encoding="utf-8", newline="")
    writer = csv.writer(wrapper)
    writer.writerow(["Name", "Amount", "Due Day", "Frequency", "Category", "Auto Pay", "Reminder Days"])
    for bill in bills:
        writer.writerow([
            bill.name,
            float(bill.amount),
            bill.due_day,
            bill.frequency,
            bill.category or "",
            "Yes" if bill.auto_pay else "No",
            bill.reminder_days,
        ])
    wrapper.flush()
    wrapper.detach()
    buf.seek(0)
    return buf


def export_debts_csv(debts: list) -> io.BytesIO:
    buf = io.BytesIO()
    wrapper = io.TextIOWrapper(buf, encoding="utf-8", newline="")
    writer = csv.writer(wrapper)
    writer.writerow(["Name", "Type", "Balance", "Credit Limit", "APR%", "Minimum Payment", "Due Day", "Auto Pay"])
    for debt in debts:
        writer.writerow([
            debt.name,
            float(debt.balance),
            debt.type,
            float(debt.credit_limit) if debt.credit_limit else "",
            float(debt.apr),
            float(debt.minimum_payment),
            debt.due_day,
            "Yes" if debt.auto_pay else "No",
        ])
    wrapper.flush()
    wrapper.detach()
    buf.seek(0)
    return buf


def export_payments_csv(payments: list) -> io.BytesIO:
    buf = io.BytesIO()
    wrapper = io.TextIOWrapper(buf, encoding="utf-8", newline="")
    writer = csv.writer(wrapper)
    writer.writerow(["Date Paid", "Bill/Debt Name", "Amount", "Pay Period", "Extra Payment"])
    for p in payments:
        name = ""
        if hasattr(p, "bill") and p.bill:
            name = p.bill.name
        elif hasattr(p, "debt") and p.debt:
            name = p.debt.name
        writer.writerow([
            str(p.paid_date) if p.paid_date else "",
            name,
            float(p.amount),
            str(p.pay_period_date) if p.pay_period_date else "",
            "Yes" if p.is_extra else "No",
        ])
    wrapper.flush()
    wrapper.detach()
    buf.seek(0)
    return buf


# ── CSV imports ─────────────────────────────────────────────────────


def import_bills_csv(file_content: bytes, user_id: UUID) -> dict:
    valid_rows: list[dict] = []
    errors: list[str] = []

    try:
        text = file_content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))

    for row_num, row in enumerate(reader, 2):
        row_errors = []

        name = (row.get("name") or row.get("Name") or "").strip()
        if not name:
            row_errors.append("name is required")

        amount_str = (row.get("amount") or row.get("Amount") or "").strip()
        amount = None
        try:
            amount = Decimal(amount_str)
            if amount <= 0:
                row_errors.append("amount must be greater than 0")
        except (InvalidOperation, ValueError):
            row_errors.append("invalid amount")

        due_day_str = (row.get("due_day") or row.get("Due Day") or "").strip()
        due_day = None
        try:
            due_day = int(due_day_str)
            if due_day < 1 or due_day > 31:
                row_errors.append("due_day must be between 1 and 31")
        except (ValueError, TypeError):
            row_errors.append("invalid due_day")

        frequency = (row.get("frequency") or row.get("Frequency") or "monthly").strip().lower()
        if frequency not in ("weekly", "biweekly", "semi_monthly", "monthly", "quarterly", "annual", "annually"):
            row_errors.append(f"invalid frequency: {frequency}")
        if frequency == "annually":
            frequency = "annual"

        category = (row.get("category") or row.get("Category") or "").strip() or None

        auto_pay_str = (row.get("auto_pay") or row.get("Auto Pay") or "false").strip().lower()
        auto_pay = auto_pay_str in ("true", "yes", "1")

        if row_errors:
            errors.append(f"Row {row_num}: {'; '.join(row_errors)}")
            continue

        valid_rows.append({
            "user_id": user_id,
            "name": name,
            "amount": amount,
            "due_day": due_day,
            "frequency": frequency,
            "category": category,
            "auto_pay": auto_pay,
        })

    return {"valid_rows": valid_rows, "errors": errors}


def import_debts_csv(file_content: bytes, user_id: UUID) -> dict:
    valid_rows: list[dict] = []
    errors: list[str] = []

    try:
        text = file_content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    valid_types = {"credit_card", "student_loan", "auto_loan", "mortgage", "personal_loan", "medical", "other"}

    for row_num, row in enumerate(reader, 2):
        row_errors = []

        name = (row.get("name") or row.get("Name") or "").strip()
        if not name:
            row_errors.append("name is required")

        debt_type = (row.get("type") or row.get("Type") or "other").strip().lower().replace(" ", "_")
        if debt_type not in valid_types:
            row_errors.append(f"invalid type: {debt_type}")

        balance = None
        balance_str = (row.get("balance") or row.get("Balance") or "").strip()
        try:
            balance = Decimal(balance_str)
            if balance < 0:
                row_errors.append("balance must be >= 0")
        except (InvalidOperation, ValueError):
            row_errors.append("invalid balance")

        credit_limit = None
        cl_str = (row.get("credit_limit") or row.get("Credit Limit") or "").strip()
        if cl_str:
            try:
                credit_limit = Decimal(cl_str)
            except (InvalidOperation, ValueError):
                row_errors.append("invalid credit_limit")

        apr = None
        apr_str = (row.get("apr") or row.get("APR%") or row.get("APR") or "").strip()
        try:
            apr = Decimal(apr_str)
            if apr < 0:
                row_errors.append("apr must be >= 0")
        except (InvalidOperation, ValueError):
            row_errors.append("invalid apr")

        minimum_payment = None
        mp_str = (row.get("minimum_payment") or row.get("Minimum Payment") or "").strip()
        try:
            minimum_payment = Decimal(mp_str)
            if minimum_payment < 0:
                row_errors.append("minimum_payment must be >= 0")
        except (InvalidOperation, ValueError):
            row_errors.append("invalid minimum_payment")

        due_day = None
        dd_str = (row.get("due_day") or row.get("Due Day") or "").strip()
        try:
            due_day = int(dd_str)
            if due_day < 1 or due_day > 31:
                row_errors.append("due_day must be between 1 and 31")
        except (ValueError, TypeError):
            row_errors.append("invalid due_day")

        if row_errors:
            errors.append(f"Row {row_num}: {'; '.join(row_errors)}")
            continue

        valid_rows.append({
            "user_id": user_id,
            "name": name,
            "type": debt_type,
            "balance": balance,
            "credit_limit": credit_limit,
            "apr": apr,
            "minimum_payment": minimum_payment,
            "due_day": due_day,
        })

    return {"valid_rows": valid_rows, "errors": errors}
